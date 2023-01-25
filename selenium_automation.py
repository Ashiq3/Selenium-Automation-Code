from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import openpyxl

# Open the Excel sheet containing the keywords
workbook = openpyxl.load_workbook('keywords.xlsx')
sheet = workbook.active

# Start the webdriver
driver = webdriver.Chrome()

# Iterate through the rows in the Excel sheet
for row in sheet.iter_rows(values_only=True):
    keyword = row[0]
    driver.get("http://www.google.com")
    search_box = driver.find_element(By.NAME,"q")
    search_box.send_keys(keyword)

    # Wait for the suggestions to load
    suggestions = driver.find_elements(By.CSS_SELECTOR,".sbtc .sbl1")
    suggestions_text = [s.text for s in suggestions]
    if len(suggestions_text) == 0:
        print(f"No suggestion found for {keyword}")
        continue
    try:
        shortest = min(suggestions_text, key=len)
        longest = max(suggestions_text, key=len)
    except ValueError:
        print(f"No suggestion found for {keyword}")
        continue
    print(shortest, longest)
    # Save the suggestions to a new Excel sheet
    sheet.cell(row=row.row, column=2).value = shortest
    sheet.cell(row=row.row, column=3).value = longest
workbook.save('result.xlsx')
# Close the browser
driver.quit()
